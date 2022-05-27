# INIT
# LOAD MODULES

from flask import Flask, render_template, request, make_response, send_file, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
#from pyexcel_io import get_data
import pandas as pd
import random
import os

# FLASK settings and INIT
host_name = "localhost"
host_port = 5000
app       = Flask(__name__)

# Upload Folder
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'csv', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Read Excel and Generate HTML Table
@app.route("/")
def homepage():
		""" Display the hyperlinks for viewing the data """
		return render_template("index.html")

@app.route("/viewBiodiversityData")
def view_BiodiversityData():
	""" Display the hyperlinks for viewing the data """
	return render_template("view_BiodiversityData.html")

@app.route("/generateReports")
def generate_Reports():
	""" Display the hyperlinks for viewing the data """
	return render_template("generate_Reports.html")

@app.route("/semanticallyIntegratedQueries")
def semantically_IntegratedQueries():
	""" Display the hyperlinks for viewing the data """
	return render_template("semantically_IntegratedQueries.html")

@app.route("/Pollinator")
def pollinator():
	""" Display the hyperlinks for viewing the data """
	return render_template("display_Pollinator.html")

@app.route("/Soil")
def soil():
	""" Display the hyperlinks for viewing the data """
	return render_template("display_Soil.html")

@app.route("/Botany")
def botany():
	""" Display the hyperlinks for viewing the data """
	return render_template("display_Botany.html", colorcode='#BDECD4')

@app.route("/SolarEnergy")
def solarenergy():
	return render_template('display_SolarEnergy.html')

@app.route("/SolarParkStudy")
def solarparkstudy():
	return render_template('display_SolarParkStudy.html')	

@app.route("/uploadNewFiles")
def upload_newFiles():
	return render_template('upload_newFiles.html')	

@app.route("/displayPollinator/<string:data>", methods=["POST", "GET"])
def displayPollinator(data):
	""" View function for displaying one aspect of the data """
	if data == 'Quadrats':
		data_df = pd.read_csv('static/PollinatorData/Field_data_2021_Quadrats_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'Quadrats_Species':
		data_df = pd.read_csv('static/PollinatorData/Pollinator_quadrat_perspecies_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == 'Species_Site_Quadrat':
		data_df = pd.read_csv('static/PollinatorData/Pollinator_quadratspecies_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == "Transects": 
		data_df = pd.read_csv('static/PollinatorData/Field_data_2021_Transects_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == "Transects_Species": 
		data_df = pd.read_csv('static/PollinatorData/Pollinator_transect_perspecies_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == "Species_Site_Transect": 
		data_df = pd.read_csv('static/PollinatorData/Pollinator_transectspecies_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == "SiteInfo": 
		data_df = pd.read_csv('static/PollinatorData/siteInfo_Anonym.csv', nrows=10)
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == "GroupSpecies": 
		data_df = pd.read_csv('static/PollinatorData/group_species.csv')
		return render_template("display_Pollinator.html", message=data, tables=[data_df.to_html()], titles=[''])


	return render_template("display_Pollinator.html")

		#data = get_data('static/PollinatorData/Field_data_2021_Quadrats_Anonym.csv')

		#book  = load_workbook('static/PollinatorData/Field_data_2021_Quadrats_Anonym.csv')  #s1_dummy.xlsx
		#sheet = book.active
		#return render_template("display.html", message=test, sheet=data)


	#if type == 'Quadrats':
	#	book  = load_workbook("static/PollinatorData/Field_data_2021_Quadrats_Anonym.csv")
	#	sheet = book.active
	#	return render_template("display.html", sheet=sheet, type=type)
	#return render_template("display.html")


@app.route("/displaySoil/<string:data>", methods=["POST", "GET"])
def displaySoil(data):
	""" View function for displaying one aspect of the data """
	if data == 'soil_data':
		data_df = pd.read_csv('static/SoilData/Soil_data_Data_Anonym_cleaneddata.csv', nrows=10)
		return render_template("display_Soil.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'soil_siteinfo':
		data_df = pd.read_csv('static/SoilData/soil_data_sitedata12_Anonym.csv', nrows=10)
		return render_template("display_Soil.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == 'soil_siteloc':
		data_df = pd.read_csv('static/SoilData/soil_data_sitedata_Anonym.csv', nrows=10)
		return render_template("display_Soil.html", message=data, tables=[data_df.to_html()], titles=[''])
	elif data == 'soil_infopersite':
		data_df = pd.read_csv('static/SoilData/soil_data_sitedata3456_Anonym.csv', nrows=10)
		return render_template("display_Soil.html", message=data, tables=[data_df.to_html()], titles=[''])		
		

@app.route("/displayBotany/<string:data>", methods=["POST", "GET"])
def displayBotany(data):
	""" View function for displaying one aspect of the data """
	if data == 'botany_data':
		book = load_workbook("static/BotanyData/SiteX_botany_1.0_Anonym.xlsx")
		#sheets = book.worksheets
		#book.active
		#sheets_dict = pd.read_excel("static/BotanyData/SiteX_botany_1.0_Anonym.xlsx")
		return render_template("display_Botany.html", message=data, spreadsheet=book, colorcode='#96A1F8 ')	

	elif data == 'botany_monitoringsurvey':
		book = load_workbook("static/BotanyData/SiteXMonitoringSurvey_Anonym.xlsx")
		#sheets = book.worksheets
		#sheets = book.sheetnames  #book.active
		#sheets_dict = pd.read_excel("static/BotanyData/SiteXMonitoringSurvey_Anonym.xlsx")
		return render_template("display_Botany.html", message=data, spreadsheet=book, colorcode='#96A1F8 ')	


@app.route("/displaySolarEnergy/<string:data>", methods=["POST", "GET"])
def displaySolarEnergy(data):
	""" View function for displaying one aspect of the data """
	if data == 'met_ghg_veg':
		book = load_workbook('static/SolarEnergyData/integrated_met_GHG_and_veg.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')	
	elif data == 'soil_and_veg':
		book = load_workbook('static/SolarEnergyData/soil_and_veg_summary.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == '2013_fulloutput':
		book = load_workbook('static/SolarEnergyData/2013_SiteY_full_output_including_nighttime.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == '2013_export':
		book = load_workbook('static/SolarEnergyData/SiteY_2013_export.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2013_temp": 
		book = load_workbook('static/SolarEnergyData/SiteY_2013_temp.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2014_export": 
		book = load_workbook('static/SolarEnergyData/SiteY_2014_export.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2014_temp": 
		book = load_workbook('static/SolarEnergyData/SiteY_2014_temp.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2015_export": 
		book = load_workbook('static/SolarEnergyData/SiteY_2015_export.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2015_temp": 
		book = load_workbook('static/SolarEnergyData/SiteY_2015_temp.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2016_export": 
		book = load_workbook('static/SolarEnergyData/SiteY_2016_export.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		
	elif data == "2016_temp": 
		book = load_workbook('static/SolarEnergyData/SiteY_2016_temp.xlsx')
		return render_template("display_SolarEnergy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')		

	return render_template("display_SolarEnergy.html")


@app.route("/displaySolarParkStudy/<string:data>", methods=["POST", "GET"])
def displaySolarParkStudy(data):
	""" View function for displaying one aspect of the data """
	if data == 'entire':
		book = load_workbook('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_Anonym.xlsx')
		return render_template("display_SolarParkStudy.html", message=data, spreadsheet=book, colorcode='#EDFAF3')	
	elif data == 'pack1':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteKI1I2I3I4II6I7_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'pack2':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteD1D2D3D4D5DLL1L2L3L4L5_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'pack3':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteL6L7L8L9L10L11L12L13M1M2MH_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'pack4':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteHHH1H2G1G2G3G4G5GN1N2N3NN4_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'pack5':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteA1AAAA2A3A4A5A6A7A8A9A10A11A12A13A14_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	
	elif data == 'pack6':
		data_df = pd.read_csv('static/SolarParkStudyData/Solar_Park_Study_Raw_Data_20200511a_anonymised_SiteFEE1E2E3E4E5E6E7CB1B2B3B4BJ_cleaned.csv', nrows=10)
		return render_template("display_SolarParkStudy.html", message=data, tables=[data_df.to_html()], titles=[''])	



#### GENERATE REPORTS
@app.route("/generateReports_Botany/", methods=["POST", "GET"])
def generateReports_botany():
	""" View function for displaying one aspect of the data """
	return render_template("generateReports_Botany.html", colorcode='#EDFAF3')

@app.route("/generateReports_SolarEnergy/", methods=["POST", "GET"])
def generateReports_solarenergy():
	""" View function for displaying one aspect of the data """
	return render_template("generateReports_SolarEnergy.html", colorcode='#EDFAF3')


@app.route("/generateReports_solarparkstudy/", methods=["POST", "GET"])
def generateReports_solarparkstudy():
	""" View function for displaying one aspect of the data """
	return render_template("generateReports_SolarParkStudy.html", colorcode='#EDFAF3')



@app.route("/displayReport_Botany/", methods=["POST", "GET"])
def displayReport_Botany():
	""" View function for displaying one aspect of the data """
	path='static/BotanyData/CreateReport_DataPrep_SiteX_botany.pdf'
	try:
		return send_file(path, attachment_filename='CreateReport_DataPrep_SiteX_botany.pdf')
	except FileNotFoundError:
		abort(404)	


@app.route("/displayReport_SolarEnergy/<string:data>", methods=["POST", "GET"])
def displayReport_SolarEnergy(data):
	""" View function for displaying one aspect of the data """

	if data == 'fulloutput':
		path     ='static/SolarEnergyData/CreateReport_DataPrep_SiteY_fulloutput.pdf.zip'
		filename = 'CreateReport_DataPrep_SiteY_fulloutput.pdf.zip'
	elif data == 'ghg':
		path     ='static/SolarEnergyData/CreateReport_DataPrep_SiteY_ghg.pdf'
		filename = 'CreateReport_DataPrep_SiteY_ghg.pdf'
	elif data == 'met':
		path     ='static/SolarEnergyData/CreateReport_DataPrep_SiteY_met.pdf'
		filename = 'CreateReport_DataPrep_SiteY_met.pdf'
	elif data == 'veg':
		path     ='static/SolarEnergyData/CreateReport_DataPrep_SiteY_veg.pdf'
		filename = 'CreateReport_DataPrep_SiteY_veg.pdf'

	try:
		return send_file(path, attachment_filename=filename)
	except FileNotFoundError:
		abort(404)


@app.route("/displayReport_SolarParkStudy/", methods=["POST", "GET"])
def displayReport_SolarParkStudy():
	""" View function for displaying one aspect of the data """
	path='static/SolarParkStudyData/CreateReport_DataPrep_SolarPark_SiteA.pdf.zip'
	try:
		return send_file(path, attachment_filename='CreateReport_DataPrep_SolarPark_SiteA.pdf.zip')
	except FileNotFoundError:
		abort(404)	

#### Semantically Integrated Queries

@app.route("/query_PollinatorSoil")
def query_Pollinator_Soil():
	""" Display the hyperlinks for viewing the data """
	return render_template("query_Pollinator_Soil.html")

@app.route("/query_PollinatorTransectQuadrat")
def query_Pollinator_Transect_Quadrat():
	""" Display the hyperlinks for viewing the data """
	return render_template("query_Pollinator_Transect_Quadrat.html")

	
@app.route("/query_PollinatorProvenance")
def query_Pollinator_Provenance():
	""" Display the hyperlinks for viewing the data """
	return render_template("query_Pollinator_Provenance.html")

@app.route("/query_PollinatorSoilSolarPark")
def query_Pollinator_Soil_SolarPark():
	""" Display the hyperlinks for viewing the data """
	return render_template("query_Pollinator_Soil_SolarPark.html")

@app.route("/query_SoilSiteInfoSolarPark")
def query_Soil_SiteInfo_SolarPark():
	""" Display the hyperlinks for viewing the data """
	return render_template("query_Pollinator_Soil_SiteInfo_SolarPark.html")

@app.route("/displayIntegratedQueries/<string:data>", methods=["POST", "GET"])
def displayIntegratedQueries(data):
	""" View function for displaying one aspect of the data """
	n = 100
	s = 10
	skip = sorted(random.sample(range(n),n-s))
    #df = pandas.read_csv(filename, skiprows=skip)


	if data == 'pollinator_soil_siteA':
		df_count = pd.read_csv('static/semanticData/pollinator_soil_siteA.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		#rowcount = sum(1 for row in data_count)
		skip = rowcount-2
		#data_df = pd.read_csv('static/semanticData/pollinator_soil_siteA.csv', nrows=10)
		
		data_df1  = pd.read_csv('static/semanticData/pollinator_soil_siteA.csv', nrows=5)
		data_df2  = pd.read_csv('static/semanticData/pollinator_soil_siteA.csv', skiprows=skip)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename','date','time','transect','transectGPS','airtemperature','windspeed','species','plotid','treatment','lat','long','soiltemp','soil_fresh_mass','soil_dry_mass','soil_moisture','soil_sample_volume','soil_bulk_density','soil_ph','soil_ammonium','soil_nitrate','agb_grasses','total_agb','agb_forbs','agb_rushes','sample_id','accuracy','plant_cover','grasses','rushes','forbs','ferns'])

		return render_template("query_Pollinator_Soil.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)	

	elif data == 'pollinator_transect_quadrat_siteB':
		df_count = pd.read_csv('static/semanticData/pollinator_transect_quadrat_siteB.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		#rowcount = sum(1 for row in data)
		skip     = rowcount-2

		data_df1 = pd.read_csv('static/semanticData/pollinator_transect_quadrat_siteB.csv', nrows=5)
		data_df2 = pd.read_csv('static/semanticData/pollinator_transect_quadrat_siteB.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['site','method','specieslist','speciescount'])
		return render_template("query_Pollinator_Transect_Quadrat.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

		
	elif data == 'pollinator_soil_siteInfo':
		df_count = pd.read_csv('static/semanticData/pollinator_soil_SiteInfo.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/pollinator_soil_SiteInfo.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/pollinator_soil_SiteInfo.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitevalue','date','obsValue','polarity','subjectivity','honeyhives','sheep','siteinfo','plots'])
		return render_template("query_Pollinator_Soil.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)
	
	elif data == 'pollinator_soil_siteInfo_siteA':
		df_count = pd.read_csv('static/semanticData/pollinator_soil_siteInfo_siteA.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/pollinator_soil_siteInfo_siteA.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/pollinator_soil_siteInfo_siteA.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['site','date','quadrat','quadratGPS','vegetationHeight','dropDisc','speciesName','percentCover','dropdisc','weightvalue','weightunit','obsValue','polarity','subjectivity','honeyhives','sheep'])
		return render_template("query_Pollinator_Soil.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

	elif data == 'pollinator_provenance':
		df_count = pd.read_csv('static/semanticData/pollinator_provenance.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/pollinator_provenance.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/pollinator_provenance.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitevalue','authorname','methodname','weight','unit'])
		return render_template("query_Pollinator_Provenance.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)


	elif data == 'pollinator_soil_solarpark_transect1':
		df_count = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteA.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteA.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteA.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename', 'comp1value',	'comp2value',	'comp3value',	'comp4value',	'comp5value',	'comp6value',	'comp7value',	'comp8value',	'comp9value',	'comp10value',	'comp11value',	'comp12value',	'comp13value',	'comp14value',	'comp15value',	'comp16value',	'comp17value',	'comp18value',	'comp19value',	'comp20value',	'comp21value',	'comp22value',	'comp23value',	'comp24value',	'comp25value',	'comp26value',	'comp27value',	'comp28value',	'comp29value',	'comp30value',	'comp31value',	'comp32value',	'comp33value',	'comp34value',	'comp35value',	'comp36value',	'comp37value',	'comp38value',	'comp39value',	'comp40value',	'comp41value',	'comp42value',	'comp43value',	'comp44value',	'comp45value',	'comp46value',	'comp47value',	'comp48value',	'comp49value',	'comp50value',	'comp51value',	'comp52value',	'comp53value',	'comp54value',	'comp55value',	'comp56value',	'comp57value',	'comp58value',	'comp59value',	'comp60value',	'comp61value',	'comp62value',	'comp63value',	'comp64value',	'date',	'time',	'transect',	'transectGPS','airtemperature','windspeed','species','plotid','treatment','lat',	'long',	'soiltemp',	'soil_fresh_mass','soil_dry_mass','soil_moisture','soil_sample_volume','soil_bulk_density','soil_ph',	'soil_ammonium','soil_nitrate',	'agb_grasses','total_agb','agb_forbs','agb_rushes','sample_id','accuracy',	'plant_cover','grasses','rushes','forbs','ferns'])
		return render_template("query_Pollinator_Soil_SolarPark.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

	elif data == 'pollinator_soil_solarpark_transect2':
		df_count = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteC.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteC.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/query-soil_PollinatorTransect_SolarPark_SiteC.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename', 'comp1value',	'comp2value',	'comp3value',	'comp4value',	'comp5value',	'comp6value',	'comp7value',	'comp8value',	'comp9value',	'comp10value',	'comp11value',	'comp12value',	'comp13value',	'comp14value',	'comp15value',	'comp16value',	'comp17value',	'comp18value',	'comp19value',	'comp20value',	'comp21value',	'comp22value',	'comp23value',	'comp24value',	'comp25value',	'comp26value',	'comp27value',	'comp28value',	'comp29value',	'comp30value',	'comp31value',	'comp32value',	'comp33value',	'comp34value',	'comp35value',	'comp36value',	'comp37value',	'comp38value',	'comp39value',	'comp40value',	'comp41value',	'comp42value',	'comp43value',	'comp44value',	'comp45value',	'comp46value',	'comp47value',	'comp48value',	'comp49value',	'comp50value',	'comp51value',	'comp52value',	'comp53value',	'comp54value',	'comp55value',	'comp56value',	'comp57value',	'comp58value',	'comp59value',	'comp60value',	'comp61value',	'comp62value',	'comp63value',	'comp64value',	'date',	'time',	'transect',	'transectGPS','airtemperature','windspeed','species','plotid','treatment','lat',	'long',	'soiltemp',	'soil_fresh_mass','soil_dry_mass','soil_moisture','soil_sample_volume','soil_bulk_density','soil_ph',	'soil_ammonium','soil_nitrate',	'agb_grasses','total_agb','agb_forbs','agb_rushes','sample_id','accuracy',	'plant_cover','grasses','rushes','forbs','ferns'])
		return render_template("query_Pollinator_Soil_SolarPark.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

	elif data == 'pollinator_soil_solarpark_quadrat':
		df_count = pd.read_csv('static/semanticData/query-soil_PollinatorQuadrat_SolarPark_SiteA.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/query-soil_PollinatorQuadrat_SolarPark_SiteA.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/query-soil_PollinatorQuadrat_SolarPark_SiteA.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename',	'comp1value',	'comp2value',	'comp3value',	'comp4value',	'comp5value',	'comp6value',	'comp7value',	'comp8value',	'comp9value',	'comp10value',	'comp11value',	'comp12value',	'comp13value',	'comp14value',	'comp15value',	'comp16value',	'comp17value',	'comp18value',	'comp19value',	'comp20value',	'comp21value',	'comp22value',	'comp23value',	'comp24value',	'comp25value',	'comp26value',	'comp27value',	'comp28value',	'comp29value',	'comp30value',	'comp31value',	'comp32value',	'comp33value',	'comp34value',	'comp35value',	'comp36value',	'comp37value',	'comp38value',	'comp39value',	'comp40value',	'comp41value',	'comp42value',	'comp43value',	'comp44value',	'comp45value',	'comp46value',	'comp47value',	'comp48value',	'comp49value',	'comp50value',	'comp51value',	'comp52value',	'comp53value',	'comp54value',	'comp55value',	'comp56value',	'comp57value',	'comp58value',	'comp59value',	'comp60value',	'comp61value',	'comp62value',	'comp63value',	'comp64value',	'date',	'quadrat',	'quadratGPS',	'vegetationHeight',	'species',	'percentCover',	'plotid',	'treatment',	'lat',	'long',	'soiltemp',	'soil_fresh_mass',	'soil_dry_mass',	'soil_moisture',	'soil_sample_volume',	'soil_bulk_density',	'soil_ph',	'soil_ammonium'	,'soil_nitrate',	'agb_grasses',	'total_agb',	'agb_forbs',	'agb_rushes',	'sample_id',	'accuracy',	'plant_cover',	'grasses',	'rushes',	'forbs',	'ferns'])
		return render_template("query_Pollinator_Soil_SolarPark.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

	elif data == 'pollinator_soil_siteinfo_solarpark1':
		df_count = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteA.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteA.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteA.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename',	'comp1value',	'comp2value',	'comp3value',	'comp4value',	'comp5value',	'comp6value',	'comp7value',	'comp8value',	'comp9value',	'comp10value',	'comp11value',	'comp12value',	'comp13value',	'comp14value',	'comp15value',	'comp16value',	'comp17value',	'comp18value',	'comp19value',	'comp20value',	'comp21value',	'comp22value',	'comp23value',	'comp24value',	'comp25value',	'comp26value',	'comp27value',	'comp28value',	'comp29value',	'comp30value',	'comp31value',	'comp32value',	'comp33value',	'comp34value',	'comp35value',	'comp36value',	'comp37value',	'comp38value',	'comp39value',	'comp40value',	'comp41value',	'comp42value',	'comp43value',	'comp44value',	'comp45value',	'comp46value',	'comp47value',	'comp48value',	'comp49value',	'comp50value',	'comp51value',	'comp52value',	'comp53value',	'comp54value',	'comp55value',	'comp56value',	'comp57value',	'comp58value',	'comp59value',	'comp60value',	'comp61value',	'comp62value',	'comp63value',	'comp64value', 'date',	'obsValue',	'polarity',	'subjectivity',	'honeyhives','sheep'])
		return render_template("query_Pollinator_Soil_SiteInfo_SolarPark.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)

	elif data == 'pollinator_soil_siteinfo_solarpark2':
		df_count = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteB.csv')
		rowcount = len(df_count.index) #sum(1 for row in data_count)
		skip     = rowcount-2  #rowcount-int(rowcount/2)

		data_df1 = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteB.csv', nrows=2)
		data_df2 = pd.read_csv('static/semanticData/query_pollinator_siteInfo_SolarPark_SiteB.csv', skiprows=skip)
		#data_df = pd.merge(data_df1, data_df2, how='outer')
		#data_df.append(data_df2, ignore_index=True)
		#data_df = pd.concat([data_df1, data_df2], axis=1)
		data_list = data_df1.values.tolist() + data_df2.values.tolist() 
		data_df   = pd.DataFrame(data_list, columns=['sitename',	'comp1value',	'comp2value',	'comp3value',	'comp4value',	'comp5value',	'comp6value',	'comp7value',	'comp8value',	'comp9value',	'comp10value',	'comp11value',	'comp12value',	'comp13value',	'comp14value',	'comp15value',	'comp16value',	'comp17value',	'comp18value',	'comp19value',	'comp20value',	'comp21value',	'comp22value',	'comp23value',	'comp24value',	'comp25value',	'comp26value',	'comp27value',	'comp28value',	'comp29value',	'comp30value',	'comp31value',	'comp32value',	'comp33value',	'comp34value',	'comp35value',	'comp36value',	'comp37value',	'comp38value',	'comp39value',	'comp40value',	'comp41value',	'comp42value',	'comp43value',	'comp44value',	'comp45value',	'comp46value',	'comp47value',	'comp48value',	'comp49value',	'comp50value',	'comp51value',	'comp52value',	'comp53value',	'comp54value',	'comp55value',	'comp56value',	'comp57value',	'comp58value',	'comp59value',	'comp60value',	'comp61value',	'comp62value',	'comp63value',	'comp64value', 'date',	'obsValue',	'polarity',	'subjectivity',	'honeyhives','sheep'])
		return render_template("query_Pollinator_Soil_SiteInfo_SolarPark.html", message=data, tables=[data_df.to_html()], titles=[''], num=skip)



	return render_template("query_Pollinator_Soil.html")

##### Uploading Files ####

def allowed_file(filename):
	return '.' in filename and \
			filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/uploadFiles/", methods=["POST", "GET"])
def upload_files():
	""" View function for displaying one aspect of the data """
	if request.method == 'POST':
		# check if the post request has the file part
		if 'file' not in request.files:
			flash('No file part')
			return redirect(request.url)
		file = request.files['file']

		# If the user does not select a file, the browser submits an empty file without a filename
		if file.filename == '':
			flash('No selected file')
			return redirect(request.url)

		if file and allowed_file(file.filename):
			filename = secure_filename(file.filename)
			file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

			#return redirect(url_for('d'))
			return render_template("uploadFiles.html", message=filename + ' has been uploaded successfully')
	return render_template("uploadFiles.html")

def files(path):
	for file in os.listdir(path):
		if os.path.isfile(os.path.join(path, file)):
			yield file

@app.route("/viewFiles/", methods=["POST", "GET"])
def view_files():
	""" View function for displaying one aspect of the data """
	path = app.config["UPLOAD_FOLDER"]
	return render_template("viewFiles.html", files=files(path))


@app.route("/viewFiles/<string:data>", methods=["POST", "GET"])
def display_uploadedfiles(data):
        '''csv'''
        if data.rsplit('.', 1)[1].lower() == 'csv':
        	data_df = pd.read_csv('static/uploads/'+ data, nrows=10)
        	return render_template("viewFiles.html", data=data, tables=[data_df.to_html()], titles=[''], ext='csv')	
        elif (data.rsplit('.', 1)[1].lower() == 'xlsx'):
        	book = load_workbook('static/uploads/' + data)
        	return render_template("viewFiles.html", data=data, spreadsheet=book, colorcode='#EDFAF3', ext='xlsx')	
        return render_template("viewFiles.html")


@app.route("/testing/<string:test>", methods=["POST", "GET"])
def testing(test):
	""" View function for displaying one aspect of the data """
	if test == 'Quadrats':
		return render_template("test.html", message=test)

@app.route("/test/<string:type1>", methods=["POST", "GET"])
def test():
	if type1 == "Transects":
		return render_template("test.html")

@app.route("/redirect_route/")
def redirect_anchor():
	#{{redirect(url_for('redirect_anchor')+'#test2')}}
	return redirect(url_for('homepage'))  #_anchor='temp'
	#return render_template("index_try.html")

# START
if __name__ == "__main__":
	port = int(os.environ.get("PORT", 5000))
	app.run(debug=True,host='0.0.0.0',port=port)